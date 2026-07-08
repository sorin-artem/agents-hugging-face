from __future__ import annotations

import base64
from io import BytesIO
import os
from pathlib import Path
import subprocess
import tempfile

from ddgs import DDGS
from ddgs.exceptions import DDGSException
from openai import OpenAI
import requests
from trafilatura import extract, fetch_url


TASK_FILE_API_URL = os.getenv(
    "TASK_FILE_API_URL",
    os.getenv("GAIA_API_URL", "https://agents-course-unit4-scoring.hf.space"),
)
MAX_PAGE_CHARS = 12000
MAX_IMAGE_BYTES = 10 * 1024 * 1024
MAX_SOURCE_CHARS = 12000
MAX_TOOL_OUTPUT_CHARS = 12000
MAX_DOCUMENT_CHARS = 20000
MAX_SPREADSHEET_ROWS = 200
MAX_SPREADSHEET_COLS = 40
PYTHON_DOCKER_IMAGE = os.getenv("PYTHON_DOCKER_IMAGE", "python:3.11-slim")


def web_search_text(query: str) -> str:
    query = query.strip()
    if not query:
        return "No search query provided."

    try:
        results = DDGS().text(query, max_results=8)
    except DDGSException as error:
        return f"Search failed: {error}"

    if not results:
        return "No search results found."

    formatted_results = []
    for index, result in enumerate(results, start=1):
        title = result.get("title", "")
        url = result.get("href", "")
        snippet = result.get("body", "")
        formatted_results.append(
            f"{index}. {title}\nURL: {url}\nSnippet: {snippet}")

    return "\n\n".join(formatted_results)


def read_url_text(url: str) -> str:
    url = url.strip()
    if not url:
        return "No URL provided."

    downloaded = fetch_url(url)
    if not downloaded:
        return f"Could not fetch URL: {url}"

    text = extract(downloaded, url=url, include_comments=False)
    if not text:
        return f"Could not extract readable text from URL: {url}"

    return text[:MAX_PAGE_CHARS]


def run_python_attachment_file(
    task_id: str,
    file_name: str = "",
    timeout_seconds: int = 5,
) -> str:
    task_id = task_id.strip()
    file_name = file_name.strip()
    if not task_id:
        return "No task_id provided."
    if file_name and Path(file_name).suffix.lower() != ".py":
        return (
            f"run_python_attachment only supports .py files. "
            f"Received {file_name!r}; use analyze_document_attachment for this file."
        )

    try:
        python_bytes, _content_type, source = _download_or_read_task_file(task_id, file_name)
    except RuntimeError as error:
        return str(error)

    if not python_bytes:
        return f"Loaded Python file for task_id {task_id}, but it was empty."

    source_code = _decode_text(python_bytes)
    timeout = max(1, min(int(timeout_seconds), 20))

    with tempfile.TemporaryDirectory(prefix="attached_python_") as temp_dir:
        script_name = Path(file_name).name or f"{task_id}.py"
        script_path = Path(temp_dir) / script_name
        script_path.write_text(source_code, encoding="utf-8")

        try:
            completed = subprocess.run(
                _docker_python_command(temp_dir, script_name),
                capture_output=True,
                text=True,
                timeout=timeout + 2,
                check=False,
            )
        except FileNotFoundError:
            return (
                f"Python source: {source}\n"
                "Execution failed: Docker CLI was not found. Install/start Docker or set up "
                "another sandbox backend.\n"
                f"Source preview:\n{source_code[:MAX_SOURCE_CHARS]}"
            )
        except subprocess.TimeoutExpired as error:
            stdout = (error.stdout or "")[:MAX_TOOL_OUTPUT_CHARS]
            stderr = (error.stderr or "")[:MAX_TOOL_OUTPUT_CHARS]
            return (
                f"Python source: {source}\n"
                f"Docker image: {PYTHON_DOCKER_IMAGE}\n"
                f"Execution timed out after {timeout} seconds.\n"
                f"Stdout:\n{stdout}\n"
                f"Stderr:\n{stderr}\n"
                f"Source preview:\n{source_code[:MAX_SOURCE_CHARS]}"
            )

    return (
        f"Python source: {source}\n"
        f"Docker image: {PYTHON_DOCKER_IMAGE}\n"
        f"Exit code: {completed.returncode}\n"
        f"Stdout:\n{completed.stdout[:MAX_TOOL_OUTPUT_CHARS]}\n"
        f"Stderr:\n{completed.stderr[:MAX_TOOL_OUTPUT_CHARS]}\n"
        f"Source preview:\n{source_code[:MAX_SOURCE_CHARS]}"
    )


def analyze_document_attachment_file(
    task_id: str,
    question: str = "",
    file_name: str = "",
) -> str:
    task_id = task_id.strip()
    question = question.strip()
    file_name = file_name.strip()
    if not task_id:
        return "No task_id provided."

    try:
        document_bytes, content_type, source = _download_or_read_task_file(task_id, file_name)
    except RuntimeError as error:
        return str(error)

    if not document_bytes:
        return f"Loaded file for task_id {task_id}, but it was empty."

    extension = _guess_file_extension(file_name, content_type, document_bytes)
    try:
        extracted = _extract_attachment_text(extension, document_bytes)
    except UnsupportedAttachmentError as error:
        return (
            f"Document source: {source}\n"
            f"File type: {extension or 'unknown'}\n"
            f"{error}"
        )
    except Exception as error:
        return f"Document analysis failed: {type(error).__name__}: {error}"

    return (
        f"Document source: {source}\n"
        f"File type: {extension or 'unknown'}\n"
        f"User question: {question or '(not provided)'}\n"
        "Extracted content:\n"
        f"{extracted}"
    )


def analyze_image_attachment_file(task_id: str, question: str, file_name: str = "") -> str:
    task_id = task_id.strip()
    question = question.strip()
    file_name = file_name.strip()
    if not task_id:
        return "No task_id provided."
    if not question:
        return "No image question provided."

    api_key = os.getenv("OPEN_ROUTER_API_KEY")
    if not api_key:
        return "Set OPEN_ROUTER_API_KEY to use image analysis."

    try:
        image_bytes, content_type, source = _download_or_read_task_file(task_id, file_name)
    except RuntimeError as error:
        return str(error)

    if not image_bytes:
        return f"Loaded file for task_id {task_id}, but it was empty."
    if len(image_bytes) > MAX_IMAGE_BYTES:
        return f"Image is too large to analyze ({len(image_bytes)} bytes)."

    mime_type = _guess_image_mime(image_bytes, content_type)
    data_url = f"data:{mime_type};base64,{base64.b64encode(image_bytes).decode('ascii')}"
    client = OpenAI(
        api_key=api_key,
        base_url="https://openrouter.ai/api/v1",
    )

    try:
        completion = client.chat.completions.create(
            model=_vision_model(),
            messages=_image_analysis_messages(question, data_url),
            temperature=0,
            max_tokens=1024,
        )
    except Exception as error:
        return f"Image analysis failed: {type(error).__name__}: {error}"

    analysis = completion.choices[0].message.content or ""
    return f"Image source: {source}\n{analysis.strip()}"


class UnsupportedAttachmentError(RuntimeError):
    pass


def _download_or_read_task_file(task_id: str, file_name: str) -> tuple[bytes, str, str]:
    hf_error: Exception | None = None
    if file_name.strip():
        try:
            hf_file = _download_from_hf_dataset(file_name)
            if hf_file:
                return hf_file
        except Exception as error:
            hf_error = error

    for candidate in _local_file_candidates(task_id, file_name):
        if candidate.is_file():
            return candidate.read_bytes(), "", str(candidate)

    try:
        response = requests.get(f"{TASK_FILE_API_URL}/files/{task_id}", timeout=60)
        response.raise_for_status()
        return response.content, response.headers.get("content-type", ""), "task file API"
    except requests.RequestException as api_error:
        if hf_error:
            raise RuntimeError(
                f"Could not download file for task_id {task_id}, no local file was found, "
                "and Hugging Face dataset download failed. "
                f"Last API error: {api_error}. HF error: {type(hf_error).__name__}: {hf_error}"
            ) from hf_error
        raise RuntimeError(
            f"Could not download file for task_id {task_id} and no local file was found. "
            f"Last API error: {api_error}"
        ) from api_error


def _download_from_hf_dataset(file_name: str) -> tuple[bytes, str, str] | None:
    file_name = file_name.strip()
    if not file_name:
        return None

    try:
        from huggingface_hub import hf_hub_download
    except ImportError:
        return None

    repo_id = os.getenv("ATTACHMENT_DATASET_REPO", "gaia-benchmark/GAIA")
    repo_prefix = os.getenv("ATTACHMENT_DATASET_PREFIX", "2023/validation").strip("/")
    dataset_filename = f"{repo_prefix}/{file_name}" if repo_prefix else file_name
    token = _load_env_token() or None
    path = hf_hub_download(
        repo_id=repo_id,
        repo_type="dataset",
        filename=dataset_filename,
        token=token,
    )
    return Path(path).read_bytes(), "", f"Hugging Face dataset: {repo_id}/{dataset_filename}"


def _load_env_token() -> str:
    try:
        from dotenv import load_dotenv
    except ImportError:
        load_dotenv = None

    if load_dotenv:
        for dotenv_path in [
            Path.cwd() / ".env",
            Path(__file__).resolve().parent.parent / ".env",
        ]:
            if dotenv_path.is_file():
                load_dotenv(dotenv_path, override=False)

    return (
        os.getenv("HF_TOKEN")
        or os.getenv("HUGGINGFACE_HUB_TOKEN")
        or os.getenv("HUGGINGFACE_TOKEN")
        or ""
    )


def _local_file_candidates(task_id: str, file_name: str) -> list[Path]:
    root_dirs = [
        Path(os.getenv("ATTACHMENT_FILES_DIR", os.getenv("GAIA_FILES_DIR", ""))),
        Path.cwd(),
        Path.cwd() / "files",
        Path.cwd() / "downloads",
        Path.cwd() / "agent_gaia" / "files",
    ]
    base_dirs = [
        candidate
        for directory in root_dirs
        if str(directory) != "."
        for candidate in [directory, directory / "2023" / "validation"]
    ]
    names = [file_name] if file_name else []
    names.extend([
        f"{task_id}{suffix}"
        for suffix in [
            ".png",
            ".jpg",
            ".jpeg",
            ".webp",
            ".gif",
            ".py",
            ".pdf",
            ".xlsx",
            ".xlsm",
            ".csv",
            ".tsv",
            ".txt",
            ".md",
            ".json",
        ]
    ])
    return [
        directory / name
        for directory in base_dirs
        if str(directory) != "."
        for name in names
        if name
    ]


def _guess_file_extension(file_name: str, content_type: str, content: bytes) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix:
        return suffix

    content_type = content_type.split(";")[0].strip().lower()
    if content_type == "application/pdf" or content.startswith(b"%PDF"):
        return ".pdf"
    if content_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    } or content.startswith(b"PK\x03\x04"):
        return ".xlsx"
    if content_type.startswith("text/"):
        return ".txt"
    return ""


def _guess_image_mime(content: bytes, content_type: str) -> str:
    content_type = content_type.split(";")[0].strip().lower()
    if content_type.startswith("image/"):
        return content_type
    if content.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if content.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if content.startswith(b"GIF87a") or content.startswith(b"GIF89a"):
        return "image/gif"
    if content.startswith(b"RIFF") and content[8:12] == b"WEBP":
        return "image/webp"
    return "image/png"


def _vision_model() -> str:
    return os.getenv(
        "OPENROUTER_VISION_MODEL",
        "google/gemini-2.5-flash",
    )


def _docker_python_command(temp_dir: str, script_name: str) -> list[str]:
    return [
        "docker",
        "run",
        "--rm",
        "--network",
        "none",
        "--memory",
        "128m",
        "--cpus",
        "1",
        "--pids-limit",
        "64",
        "--read-only",
        "--tmpfs",
        "/tmp:rw,noexec,nosuid,size=16m",
        "-v",
        f"{temp_dir}:/workspace:rw",
        "-w",
        "/workspace",
        "-e",
        "PYTHONIOENCODING=utf-8",
        "-e",
        "PYTHONDONTWRITEBYTECODE=1",
        PYTHON_DOCKER_IMAGE,
        "python",
        script_name,
    ]


def _extract_attachment_text(extension: str, document_bytes: bytes) -> str:
    if extension == ".pdf":
        return _extract_pdf_text(document_bytes)
    if extension in {".xlsx", ".xlsm"}:
        return _extract_spreadsheet_text(document_bytes)
    if extension in {".txt", ".md", ".csv", ".tsv", ".json", ".xml"}:
        return _extract_text_document(document_bytes)
    if extension in {".mp3", ".wav", ".m4a", ".flac"}:
        raise UnsupportedAttachmentError("Audio attachments are not supported yet.")
    raise UnsupportedAttachmentError("Unsupported attachment type for local document analysis.")


def _extract_pdf_text(document_bytes: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(document_bytes))
    parts = [f"PDF pages: {len(reader.pages)}"]
    for page_index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        parts.append(f"\n--- Page {page_index} ---\n{text.strip()}")
        if sum(len(part) for part in parts) >= MAX_DOCUMENT_CHARS:
            parts.append("\n[Truncated: document text limit reached]")
            break
    return "\n".join(parts)[:MAX_DOCUMENT_CHARS]


def _extract_spreadsheet_text(document_bytes: bytes) -> str:
    from openpyxl import load_workbook  # pyright: ignore[reportMissingModuleSource]

    workbook = load_workbook(
        filename=BytesIO(document_bytes),
        read_only=True,
        data_only=True,
    )
    try:
        parts = [f"Workbook sheets: {', '.join(workbook.sheetnames)}"]
        for worksheet in workbook.worksheets:
            max_row = worksheet.max_row or 0
            max_column = worksheet.max_column or 0
            parts.append(
                f"\n--- Sheet: {worksheet.title} "
                f"(max rows: {max_row or 'unknown'}, max columns: {max_column or 'unknown'}) ---"
            )
            row_count = 0
            for row in worksheet.iter_rows(
                max_row=MAX_SPREADSHEET_ROWS,
                max_col=MAX_SPREADSHEET_COLS,
                values_only=True,
            ):
                row_count += 1
                formatted = [_format_cell_value(value) for value in row]
                if any(formatted):
                    parts.append("\t".join(formatted).rstrip())
                if sum(len(part) for part in parts) >= MAX_DOCUMENT_CHARS:
                    parts.append("\n[Truncated: document text limit reached]")
                    return "\n".join(parts)[:MAX_DOCUMENT_CHARS]
            if max_row > row_count or max_column > MAX_SPREADSHEET_COLS:
                parts.append(
                    "[Truncated sheet preview: "
                    f"shown {min(row_count, MAX_SPREADSHEET_ROWS)} rows and "
                    f"{min(max_column, MAX_SPREADSHEET_COLS)} columns]"
                )
        return "\n".join(parts)[:MAX_DOCUMENT_CHARS]
    finally:
        workbook.close()


def _format_cell_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def _extract_text_document(document_bytes: bytes) -> str:
    return _decode_text(document_bytes)[:MAX_DOCUMENT_CHARS]


def _decode_text(content: bytes) -> str:
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("utf-8", errors="replace")


def _image_analysis_messages(question: str, data_url: str) -> list[dict[str, object]]:
    return [
        {
            "role": "system",
            "content": (
                "You are a visual analysis tool. Do not solve the user's task directly "
                "unless it only asks for a plain description. Return structured observations "
                "that another agent can use to answer. Be careful about visible text, labels, "
                "coordinates, orientation, counts, spatial relationships, and uncertainty. "
                "If the image is a board, chart, table, diagram, map, or puzzle, describe the "
                "layout and all visible elements precisely instead of guessing the final answer. "
                "Use plain text only, no markdown formatting."
            ),
        },
        {
            "role": "user",
            "content": [
                {
                    "type": "text",
                    "text": (
                        "User question:\n"
                        f"{question}\n\n"
                        "Return this format:\n"
                        "Visual analysis:\n"
                        "Image type: ...\n"
                        "Relevant visible text/labels: ...\n"
                        "Key objects/entities and positions: ...\n"
                        "Spatial relationships/orientation: ...\n"
                        "Details relevant to the user's question: ...\n"
                        "Uncertainties: ...\n"
                        "Do not include a final answer line unless the user only asked for image description."
                    ),
                },
                {"type": "image_url", "image_url": {"url": data_url}},
            ],
        },
    ]
